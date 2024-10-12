import json
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl.styles import PatternFill

from config import DATA_DIR_NAME, TEMP_DIR_NAME
from utils import create_temp_if_not_exist


def load_input_data(path=f"{TEMP_DIR_NAME}/input.json") -> List[Dict]:
    """
    Загружает входные данные из json файла и возвращает их
    """
    with open(path) as file:
        return json.load(file)


def create_input_data(
    path=f"{DATA_DIR_NAME}/input.xlsx", output=f"{TEMP_DIR_NAME}/input.json"
):
    """
    Загружает данные из excel файла и возвращает список
    """
    file = pd.read_excel(path, converters={"ПРОИЗВОДИТЕЛЬ": str, "КОД": str})

    data = []

    for index, row in file.iterrows():
        producer = row["ПРОИЗВОДИТЕЛЬ"]
        code = row["КОД"]

        sellers = []
        for i in range(1, 6):
            seller = row[f"ПРОДАВЕЦ{i}"]
            quantity = row[f"НАЛ.{i}"]

            if pd.isna(seller) or pd.isna(quantity):
                seller = quantity = None
            elif seller == "N/A" or quantity == "N/A":
                seller = quantity = None

            sellers.append(
                {
                    "seller": seller,
                    "quantity": quantity,
                }
            )

        data.append({"producer": producer, "code": code, "sellers": sellers})

    create_temp_if_not_exist()
    with open(output, "w") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


def _load_data() -> Tuple[List[Dict], Dict, Dict, Dict]:
    """
    Загруэет данные из json файлов и возвращает их
    """
    with open(f"{TEMP_DIR_NAME}/input.json") as file:
        input_f = json.load(file)

    with open(f"{TEMP_DIR_NAME}/vag.json") as file:
        vag = json.load(file)

    with open(f"{TEMP_DIR_NAME}/savat.json") as file:
        savat = json.load(file)

    with open(f"{TEMP_DIR_NAME}/bestparts.json") as file:
        bestparts = json.load(file)

    with open(f"{TEMP_DIR_NAME}/quattro.json") as file:
        quattro = json.load(file)

    with open(f"{TEMP_DIR_NAME}/autovag.json") as file:
        autovag = json.load(file)

    return input_f, vag, savat, bestparts, quattro, autovag


def merge_changes() -> List[Dict]:
    """
    Объединяет изменения в один файл и возвращает его
    """
    input_f, vag, savat, bestparts, quattro, autovag = _load_data()

    output = []

    for item in input_f:
        code = item["code"]
        sellers = item["sellers"]

        sellers[0] = check_seller(sellers[0], vag.get(code), "VAG UA")
        sellers[1] = check_seller(sellers[1], savat.get(code), "Savat-auto")
        sellers[2] = check_seller(sellers[2], bestparts.get(code), "Bestparts BPK")
        sellers[3] = check_seller(sellers[3], quattro.get(code), "Quatro")
        sellers[4] = check_seller(sellers[4], autovag.get(code), "AutoVag")

        item["sellers"] = sellers

        output.append(item)

    return output


def check_seller(seller, collect_item, item_name) -> Dict:
    """
    Смотрит изменения во входязем файле и возращает их
    """
    if collect_item is None:
        return seller

    if seller["seller"] is None:
        if collect_item["is_available"]:
            seller["quantity"] = f'${collect_item["quantity"]}'
            seller["seller"] = f"${item_name}"

        else:
            seller["quantity"] = "N/A"
            seller["seller"] = "N/A"

    else:
        if not collect_item["is_available"]:
            seller["quantity"] = "=0"
            seller["seller"] = item_name

        elif collect_item["is_available"] and int(collect_item["quantity"]) == 0:
            seller["quantity"] = "<0"
            seller["seller"] = item_name

        elif int(collect_item["quantity"]) < int(seller["quantity"]):
            seller["quantity"] = f'<{collect_item["quantity"]}'
            seller["seller"] = item_name

        elif int(collect_item["quantity"]) > int(seller["quantity"]):
            seller["quantity"] = f'>{collect_item["quantity"]}'
            seller["seller"] = item_name

    return seller


def collect_data_frame(output) -> pd.DataFrame:
    """
    Собирает DataFrame и возвращает его
    """
    producers = []
    codes = []
    sellers1 = []
    quantities1 = []
    sellers2 = []
    quantities2 = []
    sellers3 = []
    quantities3 = []
    sellers4 = []
    quantities4 = []
    sellers5 = []
    quantities5 = []

    for item in output:
        producers.append(item["producer"])
        codes.append(item["code"])
        sellers1.append(item["sellers"][0]["seller"])
        quantities1.append(item["sellers"][0]["quantity"])
        sellers2.append(item["sellers"][1]["seller"])
        quantities2.append(item["sellers"][1]["quantity"])
        sellers3.append(item["sellers"][2]["seller"])
        quantities3.append(item["sellers"][2]["quantity"])
        sellers4.append(item["sellers"][3]["seller"])
        quantities4.append(item["sellers"][3]["quantity"])
        sellers5.append(item["sellers"][4]["seller"])
        quantities5.append(item["sellers"][4]["quantity"])

    df_output = pd.DataFrame(
        {
            "ПРОИЗВОДИТЕЛЬ": producers,
            "КОД": codes,
            "ПРОДАВЕЦ1": sellers1,
            "НАЛ.1": quantities1,
            "ПРОДАВЕЦ2": sellers2,
            "НАЛ.2": quantities2,
            "ПРОДАВЕЦ3": sellers3,
            "НАЛ.3": quantities3,
            "ПРОДАВЕЦ4": sellers4,
            "НАЛ.4": quantities4,
            "ПРОДАВЕЦ5": sellers5,
            "НАЛ.5": quantities5,
        }
    )

    return df_output


def sheet_styler(sheet, sheet_len):
    """
    Стилизирует страницу
    """
    for cell_index in range(2, sheet_len + 2):
        for cell_name in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            for (cell,) in sheet[f"{cell_name}{cell_index}:{cell_name}{cell_index}"]:
                cell_value = str(cell.value)

                if "$" in cell_value:
                    cell.value = cell_value.replace("$", "")
                    cell.fill = PatternFill("solid", start_color="FEFF54")

                elif "=" in cell_value:
                    cell.value = cell_value.replace("=", "")
                    cell.fill = PatternFill("solid", start_color="92D050")

                elif "<" in cell_value:
                    cell.value = cell_value.replace("<", "")
                    cell.fill = PatternFill("solid", start_color="F5C242")

                elif ">" in cell_value:
                    cell.value = cell_value.replace(">", "")
                    cell.fill = PatternFill("solid", start_color="95B3DE")


def excel_justify(worksheet):
    """
    Выравнивает поля в Excel файле по ширине
    """
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (length + 2) * 1.2
        worksheet.column_dimensions[column_cells[0].column_letter].width = (
            adjusted_width
        )


def collect_excel_file():
    """
    Собирает Excel файл
    """
    output = merge_changes()

    new_file_name = (
        f'{DATA_DIR_NAME}/output_{datetime.now().strftime("%d.%m.%Y(%H.%M)")}-5.xlsx'
    )

    with pd.ExcelWriter(new_file_name, engine="openpyxl") as writer:
        df_output = collect_data_frame(output)

        df_output.to_excel(writer, sheet_name="Sheet1", index=False)

        worksheet = writer.sheets["Sheet1"]

        sheet_styler(worksheet, len(output))
        excel_justify(worksheet)
